from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import load_workbook

from md_generator.ppt.vendor_pdf_md.convert import pdf_to_markdown
from md_generator.ppt.vendor_word_md.convert import docx_to_markdown_bundle

INLINE_CAP = 48_000

_EXCLUDE_PARTS = ("extracted_md",)


def _skip_asset_path(p: Path, assets_root: Path) -> bool:
    try:
        rel = p.resolve().relative_to(assets_root.resolve())
    except ValueError:
        return True
    parts = rel.parts
    return any(x in parts for x in _EXCLUDE_PARTS)


def _iter_asset_files(assets_root: Path, suffixes: tuple[str, ...]) -> list[Path]:
    out: list[Path] = []
    for p in assets_root.rglob("*"):
        if not p.is_file():
            continue
        if _skip_asset_path(p, assets_root):
            continue
        if p.suffix.lower() in suffixes:
            out.append(p)
    return sorted(out)


def _slug(s: str) -> str:
    s = re.sub(r"[^\w\-]+", "_", s, flags=re.UNICODE).strip("_")
    return s[:80] or "file"


def process_txt_assets(document_path: Path, assets_root: Path, options) -> str:
    if not options.emit_extracted_txt_md:
        return ""
    extracted_md = assets_root / "extracted_md"
    extracted_md.mkdir(parents=True, exist_ok=True)
    blocks: list[str] = []
    n = 0
    for txt in _iter_asset_files(assets_root, (".txt",)):
        n += 1
        body = txt.read_text(encoding="utf-8", errors="replace")
        out_md = extracted_md / f"{n:04d}_{_slug(txt.stem)}.md"
        out_md.write_text(f"# {txt.name}\n\n{body}", encoding="utf-8")
        rel = Path("assets") / out_md.relative_to(assets_root.parent).as_posix()
        if len(body) <= INLINE_CAP:
            blocks.append(f"### `{txt.name}`\n\n{body}\n")
        else:
            blocks.append(f"### `{txt.name}`\n\n[View extracted file]({rel})\n")
    if not blocks:
        return ""
    return "## Extracted text from package assets\n\n" + "\n".join(blocks)


def process_docx_pdf_xlsx(document_path: Path, assets_root: Path, options) -> str:
    extracted_md = assets_root / "extracted_md"
    extracted_md.mkdir(parents=True, exist_ok=True)
    media_root = assets_root / "extracted_md_media"
    media_root.mkdir(parents=True, exist_ok=True)

    blocks_word: list[str] = []
    blocks_xlsx: list[str] = []

    n = 0
    for docx in _iter_asset_files(assets_root, (".docx",)):
        if not options.extracted_docx_md:
            continue
        n += 1
        slug = f"{n:04d}_{_slug(docx.stem)}"
        media_dir = media_root / slug
        md_out = extracted_md / f"{slug}.md"
        data = docx.read_bytes()
        md, _imgs = docx_to_markdown_bundle(
            data,
            media_dir=media_dir,
            md_output_path=md_out,
        )
        md_out.write_text(md, encoding="utf-8")
        rel = Path("assets") / md_out.relative_to(assets_root.parent).as_posix()
        snippet = md if len(md) <= INLINE_CAP else f"[View extracted file]({rel})"
        blocks_word.append(f"### `{docx.name}`\n\n{snippet}\n")

    n = 0
    for pdf in _iter_asset_files(assets_root, (".pdf",)):
        if not options.extracted_pdf_md:
            continue
        n += 1
        slug = f"{n:04d}_{_slug(pdf.stem)}"
        media_dir = media_root / slug
        md_out = extracted_md / f"{slug}.md"
        data = pdf.read_bytes()
        md = pdf_to_markdown(
            data,
            media_dir=media_dir,
            ocr=options.extracted_pdf_ocr,
            ocr_min_chars=options.extracted_pdf_ocr_min_chars,
        )
        md_out.write_text(md, encoding="utf-8")
        rel = Path("assets") / md_out.relative_to(assets_root.parent).as_posix()
        snippet = md if len(md) <= INLINE_CAP else f"[View extracted file]({rel})"
        blocks_word.append(f"### `{pdf.name}`\n\n{snippet}\n")

    n = 0
    for xlsx in _iter_asset_files(assets_root, (".xlsx", ".xlsm")):
        if not options.extracted_xlsx_md:
            continue
        n += 1
        slug = f"{n:04d}_{_slug(xlsx.stem)}"
        md_out = extracted_md / f"{slug}.md"
        md = _xlsx_to_gfm_markdown(xlsx)
        md_out.write_text(md, encoding="utf-8")
        rel = Path("assets") / md_out.relative_to(assets_root.parent).as_posix()
        snippet = md if len(md) <= INLINE_CAP else f"[View extracted file]({rel})"
        blocks_xlsx.append(f"### `{xlsx.name}`\n\n{snippet}\n")

    extra = ""
    if blocks_word:
        extra += "## Extracted Word and PDF (Markdown)\n\n" + "\n".join(blocks_word)
    if blocks_xlsx:
        extra += "\n## Extracted Excel (Markdown)\n\n" + "\n".join(blocks_xlsx)
    return extra


def _xlsx_to_gfm_markdown(path: Path) -> str:
    wb = load_workbook(path, read_only=True, data_only=True)
    parts: list[str] = []
    for sheet in wb.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        parts.append(f"## {sheet.title}\n\n")
        parts.append(_rows_to_gfm(rows))
    wb.close()
    return "\n".join(parts)


def _rows_to_gfm(rows: list[tuple[object | None, ...]]) -> str:
    str_rows: list[list[str]] = []
    for row in rows:
        str_rows.append([(str(c) if c is not None else "").replace("|", "\\|") for c in row])
    if not str_rows:
        return ""
    width = max(len(r) for r in str_rows)
    norm = [r + [""] * (width - len(r)) for r in str_rows]
    header = norm[0]
    sep = ["---"] * width
    lines = [
        "| " + " | ".join(header) + " |",
        "| " + " | ".join(sep) + " |",
    ]
    for row in norm[1:]:
        lines.append("| " + " | ".join(row) + " |")
    return "\n".join(lines) + "\n"


def append_post_sections(document_path: Path, assets_root: Path, options) -> None:
    """Append post-process sections to document.md (artifact layout)."""
    doc = document_path.read_text(encoding="utf-8")
    additions = process_txt_assets(document_path, assets_root, options)
    additions += process_docx_pdf_xlsx(document_path, assets_root, options)
    if additions.strip():
        document_path.write_text(doc.rstrip() + "\n\n" + additions.strip() + "\n", encoding="utf-8")
