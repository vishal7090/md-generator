from __future__ import annotations

import csv
import logging
from pathlib import Path
from typing import Any, Iterator

from openpyxl import load_workbook

logger = logging.getLogger(__name__)


def _sheet_visible(sheet: Any, include_hidden: bool) -> bool:
    if include_hidden:
        return True
    state = getattr(sheet, "sheet_state", None) or "visible"
    return state == "visible"


def _normalize_readonly_rows(raw_rows: list[tuple[Any, ...]], max_rows: int) -> list[list[Any]]:
    out: list[list[Any]] = []
    max_col = 0
    for row in raw_rows[:max_rows]:
        lst = list(row)
        max_col = max(max_col, len(lst))
        out.append(lst)
    for r in out:
        while len(r) < max_col:
            r.append(None)
    return out


def iter_sheets_streaming(
    path: Path,
    *,
    include_hidden_sheets: bool,
    sheet_name_filter: set[str] | None,
    max_rows_per_sheet: int,
) -> Iterator[tuple[str, list[list[Any]]]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet in wb:
            if not _sheet_visible(sheet, include_hidden_sheets):
                continue
            if sheet_name_filter is not None and sheet.title.strip().casefold() not in sheet_name_filter:
                continue
            raw: list[tuple[Any, ...]] = []
            count = 0
            for row in sheet.iter_rows(values_only=True):
                if count >= max_rows_per_sheet:
                    break
                raw.append(row)
                count += 1
            yield sheet.title, _normalize_readonly_rows(raw, max_rows_per_sheet)
    finally:
        wb.close()


def _read_sheet_matrix(
    sheet: Any,
    *,
    expand_merged: bool,
    max_rows: int,
) -> list[list[Any]]:
    max_r = sheet.max_row or 0
    max_c = sheet.max_column or 0
    if max_r == 0 or max_c == 0:
        return []
    max_r = min(max_r, max_rows)
    grid: list[list[Any]] = [[None for _ in range(max_c)] for _ in range(max_r)]
    for row in sheet.iter_rows(min_row=1, max_row=max_r, min_col=1, max_col=max_c):
        for cell in row:
            grid[cell.row - 1][cell.column - 1] = cell.value
    if expand_merged:
        for m in list(sheet.merged_cells.ranges):
            min_r, min_c = m.min_row, m.min_col
            max_r2, max_c2 = m.max_row, m.max_col
            if min_r > max_r:
                continue
            max_r2 = min(max_r2, max_r)
            val = grid[min_r - 1][min_c - 1]
            for r in range(min_r - 1, max_r2):
                for c in range(min_c - 1, max_c2):
                    if c < max_c:
                        grid[r][c] = val
    return grid


def iter_sheets_buffered(
    path: Path,
    *,
    include_hidden_sheets: bool,
    sheet_name_filter: set[str] | None,
    max_rows_per_sheet: int,
    expand_merged_cells: bool,
) -> Iterator[tuple[str, list[list[Any]]]]:
    wb = load_workbook(path, read_only=False, data_only=True)
    try:
        for sheet in wb:
            if not _sheet_visible(sheet, include_hidden_sheets):
                continue
            if sheet_name_filter is not None and sheet.title.strip().casefold() not in sheet_name_filter:
                continue
            matrix = _read_sheet_matrix(
                sheet,
                expand_merged=expand_merged_cells,
                max_rows=max_rows_per_sheet,
            )
            yield sheet.title, matrix
    finally:
        wb.close()


def _read_csv_matrix(path: Path, max_rows: int) -> list[list[Any]]:
    rows: list[list[Any]] = []
    with path.open(newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        for i, row in enumerate(reader):
            if i >= max_rows:
                break
            rows.append(list(row))
    if not rows:
        return []
    max_c = max(len(r) for r in rows)
    for r in rows:
        while len(r) < max_c:
            r.append("")
    return rows


def iter_sheet_rows(
    path: Path,
    *,
    streaming: bool,
    expand_merged_cells: bool,
    include_hidden_sheets: bool,
    sheet_names: list[str] | None,
    max_rows_per_sheet: int,
    warnings: list[str],
) -> Iterator[tuple[str, list[list[Any]]]]:
    filt: set[str] | None = None
    if sheet_names:
        filt = {n.strip().casefold() for n in sheet_names}

    if path.suffix.casefold() == ".csv":
        title = path.stem
        if filt is not None and title.strip().casefold() not in filt:
            return
        matrix = _read_csv_matrix(path, max_rows_per_sheet)
        yield title, matrix
        return

    if streaming:
        if expand_merged_cells:
            warnings.append("streaming mode is active: merged cells are not expanded")
            logger.warning("streaming mode: merged-cell expansion disabled")
        yield from iter_sheets_streaming(
            path,
            include_hidden_sheets=include_hidden_sheets,
            sheet_name_filter=filt,
            max_rows_per_sheet=max_rows_per_sheet,
        )
    else:
        yield from iter_sheets_buffered(
            path,
            include_hidden_sheets=include_hidden_sheets,
            sheet_name_filter=filt,
            max_rows_per_sheet=max_rows_per_sheet,
            expand_merged_cells=expand_merged_cells,
        )
