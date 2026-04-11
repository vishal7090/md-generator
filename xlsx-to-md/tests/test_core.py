from __future__ import annotations

import io
import zipfile
from pathlib import Path

import pytest
from openpyxl import Workbook

from src.convert_config import ConvertConfig
from src.converter_core import convert_excel_to_markdown


def _write_minimal_wb(path: Path, hidden_second: bool = False, merged: bool = False) -> None:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Alpha"
    ws.append(["ID", "Name"])
    ws.append([1, "Ann"])
    ws2 = wb.create_sheet("Beta")
    ws2.append(["X"])
    ws2.append([2])
    if hidden_second:
        ws2.sheet_state = "hidden"
    if merged:
        ws3 = wb.create_sheet("Merged")
        ws3["A1"] = "Title"
        ws3.merge_cells("A1:B2")
    wb.save(path)


def test_combined_output_toc(tmp_path: Path) -> None:
    p = tmp_path / "w.xlsx"
    _write_minimal_wb(p)
    out = tmp_path / "out"
    r = convert_excel_to_markdown(p, out, config=ConvertConfig(include_toc=True, split_by_sheet=False))
    assert "Alpha" in r.sheets_processed and "Beta" in r.sheets_processed
    md = (out / "w.md").read_text(encoding="utf-8")
    assert "Table of contents" in md
    assert '<h2 id="alpha">Sheet: Alpha</h2>' in md
    assert '<h2 id="beta">Sheet: Beta</h2>' in md


def test_combined_no_toc_single_sheet(tmp_path: Path) -> None:
    p = tmp_path / "w.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(["A"])
    wb.save(p)
    out = tmp_path / "out"
    r = convert_excel_to_markdown(p, out, config=ConvertConfig(include_toc=True))
    md = (out / "w.md").read_text(encoding="utf-8")
    assert "Table of contents" not in md


def test_split_by_sheet(tmp_path: Path) -> None:
    p = tmp_path / "w.xlsx"
    _write_minimal_wb(p)
    out = tmp_path / "out"
    convert_excel_to_markdown(p, out, config=ConvertConfig(split_by_sheet=True))
    assert (out / "alpha.md").is_file()
    assert (out / "beta.md").is_file()


def test_hidden_sheet_excluded(tmp_path: Path) -> None:
    p = tmp_path / "w.xlsx"
    _write_minimal_wb(p, hidden_second=True)
    out = tmp_path / "out"
    r = convert_excel_to_markdown(p, out, config=ConvertConfig())
    assert r.sheets_processed == ["Alpha"]


def test_hidden_sheet_included(tmp_path: Path) -> None:
    p = tmp_path / "w.xlsx"
    _write_minimal_wb(p, hidden_second=True)
    out = tmp_path / "out"
    r = convert_excel_to_markdown(
        p,
        out,
        config=ConvertConfig(include_hidden_sheets=True),
    )
    assert "Beta" in r.sheets_processed


def test_sheet_name_filter(tmp_path: Path) -> None:
    p = tmp_path / "w.xlsx"
    _write_minimal_wb(p)
    out = tmp_path / "out"
    r = convert_excel_to_markdown(
        p,
        out,
        config=ConvertConfig(sheet_names=["beta"]),
    )
    assert r.sheets_processed == ["Beta"]


def test_merged_cells_expand(tmp_path: Path) -> None:
    p = tmp_path / "w.xlsx"
    _write_minimal_wb(p, merged=True)
    out = tmp_path / "out"
    r = convert_excel_to_markdown(
        p,
        out,
        config=ConvertConfig(sheet_names=["Merged"], expand_merged_cells=True, streaming=False),
    )
    assert r.sheets_processed == ["Merged"]
    md = (out / "w.md").read_text(encoding="utf-8")
    assert "Title" in md


def test_streaming_warns_no_merge_expand(tmp_path: Path) -> None:
    p = tmp_path / "w.xlsx"
    _write_minimal_wb(p, merged=True)
    out = tmp_path / "out"
    r = convert_excel_to_markdown(
        p,
        out,
        config=ConvertConfig(
            sheet_names=["Merged"],
            streaming=True,
            expand_merged_cells=True,
        ),
    )
    assert any("streaming" in w.lower() for w in r.warnings)


def test_column_indices(tmp_path: Path) -> None:
    p = tmp_path / "w.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(["A", "B", "C"])
    ws.append([1, 2, 3])
    wb.save(p)
    out = tmp_path / "out"
    convert_excel_to_markdown(p, out, config=ConvertConfig(column_indices=[0, 2]))
    md = (out / "w.md").read_text(encoding="utf-8")
    assert "| A | C |" in md
    assert "2" not in md.split("\n")[4]  # middle column dropped from body row


def test_column_names(tmp_path: Path) -> None:
    p = tmp_path / "w.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(["Foo", "Bar"])
    ws.append([10, 20])
    wb.save(p)
    out = tmp_path / "out"
    convert_excel_to_markdown(p, out, config=ConvertConfig(column_names=["Bar"]))
    md = (out / "w.md").read_text(encoding="utf-8")
    assert "| Bar |" in md
    assert "20" in md


def test_config_from_json(tmp_path: Path) -> None:
    p = tmp_path / "w.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(["x"])
    wb.save(p)
    cfg_path = tmp_path / "c.json"
    cfg_path.write_text('{"split_by_sheet": true, "include_toc": false}', encoding="utf-8")
    out = tmp_path / "out"
    cfg = ConvertConfig.from_json(cfg_path)
    convert_excel_to_markdown(p, out, config=cfg)
    assert (out / "sheet.md").is_file() or list(out.glob("*.md"))


def test_csv_input(tmp_path: Path) -> None:
    p = tmp_path / "data.csv"
    p.write_text("a,b\n1,2\n", encoding="utf-8")
    out = tmp_path / "out"
    r = convert_excel_to_markdown(p, out, config=ConvertConfig())
    assert r.sheets_processed == ["data"]
    md = (out / "data.md").read_text(encoding="utf-8")
    assert "| a | b |" in md


def test_invalid_extension(tmp_path: Path) -> None:
    p = tmp_path / "x.txt"
    p.write_text("nope", encoding="utf-8")
    with pytest.raises(ValueError, match="xlsx"):
        convert_excel_to_markdown(p, tmp_path / "out")


def test_output_basename(tmp_path: Path) -> None:
    p = tmp_path / "w.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(["a"])
    wb.save(p)
    out = tmp_path / "out"
    convert_excel_to_markdown(p, out, config=ConvertConfig(output_basename="report.md"))
    assert (out / "report.md").is_file()


def test_h3_heading_level(tmp_path: Path) -> None:
    p = tmp_path / "w.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(["a"])
    wb.save(p)
    out = tmp_path / "out"
    convert_excel_to_markdown(p, out, config=ConvertConfig(sheet_heading_level="###"))
    md = (out / "w.md").read_text(encoding="utf-8")
    assert "<h3 id=" in md


def test_alignment_row(tmp_path: Path) -> None:
    p = tmp_path / "w.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(["L", "R"])
    ws.append(["a", "b"])
    wb.save(p)
    out = tmp_path / "out"
    convert_excel_to_markdown(
        p,
        out,
        config=ConvertConfig(
            enable_alignment_in_tables=True,
            column_alignment=["left", "right"],
        ),
    )
    md = (out / "w.md").read_text(encoding="utf-8")
    assert ":---" in md and "---:" in md
